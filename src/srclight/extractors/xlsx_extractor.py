"""XLSX extractor (requires openpyxl)."""

from __future__ import annotations

import io
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from .base import make_section

if TYPE_CHECKING:
    from ..db import Database

MAX_ROWS = 500


class XlsxExtractor:
    language = "xlsx"
    extensions = (".xlsx", ".xlsm")

    def extract(
        self,
        file_id: int,
        rel_path: str,
        source: bytes,
        db: Database,
    ) -> int:
        stem = Path(rel_path).stem
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)

        count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(cells)
                if len(rows) > MAX_ROWS:
                    break

            if not rows:
                continue

            # Header row as signature
            header = rows[0]
            signature = "\t".join(header)

            # Content: all rows tab-delimited
            content_lines = ["\t".join(row) for row in rows[:MAX_ROWS]]
            content = "\n".join(content_lines)

            # Stats
            total_rows = ws.max_row or len(rows)
            col_count = ws.max_column or len(header)
            truncated = total_rows > MAX_ROWS

            col_names = ", ".join(h for h in header if h.strip())
            meta_parts = [f"{total_rows} rows", f"{col_count} columns"]
            if truncated:
                meta_parts.append(f"(truncated to {MAX_ROWS} rows)")
            stats = ", ".join(meta_parts)
            doc_comment = f"Columns: {col_names}. {stats}" if col_names else stats

            sym = make_section(
                file_id, rel_path,
                name=sheet_name,
                qualified_name=f"{stem} > {sheet_name}",
                content=content,
                signature=signature,
                doc_comment=doc_comment,
            )
            db.insert_symbol(sym, rel_path)
            count += 1

        wb.close()
        return count
